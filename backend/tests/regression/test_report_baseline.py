from __future__ import annotations

import io
import json
from pathlib import Path

import openpyxl
import pytest

from models import SizingInput
from services.report_service import ReportGenerator


def test_report_service_outputs_xlsx_and_expected_cells() -> None:
    payload_path = Path(__file__).parent.parent / "payload.json"
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    sizing_input = SizingInput(**payload)

    generator = ReportGenerator()
    result = generator.generate(sizing_input)

    assert isinstance(result, io.BytesIO)
    assert result.getbuffer().nbytes > 0
    assert result.getvalue().startswith(b"PK")

    workbook = openpyxl.load_workbook(io.BytesIO(result.getvalue()))
    worksheet = workbook.active

    assert worksheet["D4"].value == payload["internal_users"]
    assert worksheet["D23"].value == payload["params_billions"]
    assert worksheet["D38"].value == payload["gpu_mem_gb"]
    assert worksheet["D58"].value == payload["gpu_flops_Fcount"]


def test_report_service_raises_when_template_missing(tmp_path: Path) -> None:
    payload_path = Path(__file__).parent.parent / "payload.json"
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    sizing_input = SizingInput(**payload)

    generator = ReportGenerator(template_path=str(tmp_path / "missing.xlsx"))
    try:
        generator.generate(sizing_input)
        raise AssertionError("expected FileNotFoundError")
    except FileNotFoundError:
        pass


def test_report_service_uses_lookup_when_tflops_not_set(monkeypatch) -> None:
    payload_path = Path(__file__).parent.parent / "payload.json"
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    payload["gpu_flops_Fcount"] = None
    sizing_input = SizingInput(**payload)

    called = {"value": False}

    def _lookup(gpu_id, gpu_mem_gb):  # noqa: ANN001
        called["value"] = True
        assert gpu_mem_gb == payload["gpu_mem_gb"]
        return 999.0

    generator = ReportGenerator(gpu_tflops_lookup=_lookup)
    result = generator.generate(sizing_input)
    assert called["value"] is True

    workbook = openpyxl.load_workbook(io.BytesIO(result.getvalue()))
    worksheet = workbook.active
    assert worksheet["D58"].value == 999.0


def test_report_service_sets_zero_tflops_when_lookup_unavailable() -> None:
    payload_path = Path(__file__).parent.parent / "payload.json"
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    payload["gpu_flops_Fcount"] = None
    sizing_input = SizingInput(**payload)

    generator = ReportGenerator(gpu_tflops_lookup=lambda gpu_id, gpu_mem_gb: None)  # noqa: ARG005
    result = generator.generate(sizing_input)

    workbook = openpyxl.load_workbook(io.BytesIO(result.getvalue()))
    worksheet = workbook.active
    assert worksheet["D58"].value == 0


def test_report_service_wraps_template_fill_errors(monkeypatch) -> None:
    payload_path = Path(__file__).parent.parent / "payload.json"
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    sizing_input = SizingInput(**payload)

    generator = ReportGenerator()

    def _boom(inp: SizingInput):  # noqa: ARG001
        raise ValueError("broken template")

    monkeypatch.setattr(generator, "_fill_template", _boom)
    with pytest.raises(RuntimeError):
        generator.generate(sizing_input)
