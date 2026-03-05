"""
Генератор Excel-отчётов по шаблону reportTemplate.xlsx.

Заполняет ячейки шаблона, помеченные как «Заполняется», входными значениями
из SizingInput. Формулы («Рассчитывается») сохраняются и пересчитываются
Excel при открытии файла.
"""
from __future__ import annotations

import io
import os
import logging
from datetime import datetime
from typing import Callable, Optional

import openpyxl

from models import SizingInput

logger = logging.getLogger("sizing.report")

# Путь к шаблону Excel (лежит рядом с этим модулем)
REPORT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "reportTemplate.xlsx")


class ReportGenerator:
    """
    Генератор Excel-отчётов на основе шаблона.

    Принимает входные параметры расчёта (SizingInput) и опционально
    функцию для поиска TFLOPS GPU из каталога.
    """

    def __init__(
        self,
        template_path: str = REPORT_TEMPLATE_PATH,
        gpu_tflops_lookup: Optional[Callable] = None,
    ):
        self.template_path = template_path
        self._gpu_tflops_lookup = gpu_tflops_lookup

    @property
    def template_exists(self) -> bool:
        """Проверяет наличие файла шаблона."""
        return os.path.exists(self.template_path)

    def _resolve_gpu_tflops(self, inp: SizingInput) -> float:
        """
        Определяет значение TFLOPS GPU.

        Приоритет:
        1. Явно заданное значение во входных данных (gpu_flops_Fcount)
        2. Поиск в каталоге GPU по gpu_id / gpu_mem_gb
        3. 0 (если ничего не найдено)
        """
        if inp.gpu_flops_Fcount is not None:
            return inp.gpu_flops_Fcount

        if self._gpu_tflops_lookup is not None:
            tflops = self._gpu_tflops_lookup(inp.gpu_id, inp.gpu_mem_gb)
            if tflops:
                return tflops

        return 0

    def _fill_template(self, inp: SizingInput) -> openpyxl.Workbook:
        """
        Загружает шаблон и заполняет входные ячейки значениями из SizingInput.

        Возвращает готовый Workbook.
        """
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active

        # ── Раздел 2.1: Пользователи и нагрузка ──
        ws["D4"] = inp.internal_users
        ws["D5"] = inp.penetration_internal
        ws["D6"] = inp.concurrency_internal
        ws["D7"] = inp.sessions_per_user_J

        ws["E4"] = inp.external_users
        ws["E5"] = inp.penetration_external
        ws["E6"] = inp.concurrency_external
        ws["E7"] = inp.sessions_per_user_J

        for col in ("F", "G"):
            ws[f"{col}4"] = 0
            ws[f"{col}5"] = 0
            ws[f"{col}6"] = 0
            ws[f"{col}7"] = 0

        # ── Раздел 2.2: Токены ──
        ws["D13"] = inp.system_prompt_tokens_SP
        ws["D14"] = inp.user_prompt_tokens_Prp
        ws["D15"] = inp.answer_tokens_A
        ws["D16"] = inp.reasoning_tokens_MRT
        ws["D17"] = inp.dialog_turns
        ws["D19"] = inp.max_context_window_TSmax

        # ── Раздел 3.1: Модель ──
        ws["D23"] = inp.params_billions
        ws["D24"] = inp.bytes_per_param
        ws["D25"] = inp.safe_margin
        ws["D26"] = inp.emp_model

        # ── Раздел 3.2: KV-кэш ──
        ws["D31"] = inp.layers_L
        ws["D32"] = inp.hidden_size_H
        ws["D33"] = inp.bytes_per_kv_state
        ws["D34"] = inp.emp_kv

        # ── Раздел 4: GPU и Tensor Parallelism ──
        ws["D38"] = inp.gpu_mem_gb
        ws["D39"] = inp.kavail
        ws["D41"] = inp.gpus_per_server
        ws["D42"] = inp.tp_multiplier_Z
        ws["D48"] = inp.saturation_coeff_C

        # ── Раздел 6: Compute ──
        ws["D58"] = self._resolve_gpu_tflops(inp)
        ws["D60"] = inp.eta_prefill
        ws["D61"] = inp.eta_decode
        ws["D64"] = inp.th_prefill_empir if inp.th_prefill_empir else 0
        ws["D65"] = inp.th_decode_empir if inp.th_decode_empir else 0
        ws["D68"] = inp.rps_per_session_R
        ws["D69"] = inp.sla_reserve_KSLA

        # ── Раздел 7: SLA targets ──
        ws["D76"] = inp.ttft_sla if inp.ttft_sla else 0
        ws["D77"] = inp.e2e_latency_sla if inp.e2e_latency_sla else 0

        return wb

    def generate(self, inp: SizingInput) -> io.BytesIO:
        """
        Генерирует заполненный Excel-файл и возвращает его как BytesIO-буфер.

        Raises:
            FileNotFoundError: если шаблон не найден.
            RuntimeError: при ошибке генерации.
        """
        if not self.template_exists:
            raise FileNotFoundError(
                f"Шаблон отчёта не найден: {self.template_path}"
            )

        try:
            wb = self._fill_template(inp)
        except Exception as exc:
            logger.error("Ошибка при заполнении шаблона: %s", exc)
            raise RuntimeError(f"Ошибка генерации отчёта: {exc}") from exc

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def make_filename() -> str:
        """Формирует имя файла с текущей датой/временем."""
        return f"sizing_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
