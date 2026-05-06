"""
Генератор Excel-отчётов по шаблону reportTemplate.xlsx.

Заполняет ячейки шаблона, помеченные как «Заполняется», входными значениями
из SizingInput. Формулы («Рассчитывается») сохраняются и пересчитываются
Excel при открытии файла.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import openpyxl

from models import SizingInput

logger = logging.getLogger("sizing.report")

# Путь к шаблону Excel в корне backend.
REPORT_TEMPLATE_PATH = str(Path(__file__).resolve().parents[1] / "reportTemplate.xlsx")


class ReportGenerator:
    """
    Генератор Excel-отчётов на основе шаблона.

    Принимает входные параметры расчёта (SizingInput) и опционально
    функцию для поиска TFLOPS GPU из каталога.
    """

    def __init__(
        self,
        template_path: str = REPORT_TEMPLATE_PATH,
        gpu_tflops_lookup: Optional[Callable] = None,
    ):
        self.template_path = template_path
        self._gpu_tflops_lookup = gpu_tflops_lookup

    @property
    def template_exists(self) -> bool:
        """Проверяет наличие файла шаблона."""
        return Path(self.template_path).exists()

    def _resolve_gpu_tflops(self, inp: SizingInput) -> float:
        """
        Определяет значение TFLOPS GPU.

        Приоритет:
        1. Явно заданное значение во входных данных (gpu_flops_Fcount)
        2. Поиск в каталоге GPU по gpu_id / gpu_mem_gb
        3. 0 (если ничего не найдено)
        """
        if inp.gpu_flops_Fcount is not None:
            return inp.gpu_flops_Fcount

        if self._gpu_tflops_lookup is not None:
            tflops = self._gpu_tflops_lookup(inp.gpu_id, inp.gpu_mem_gb)
            if tflops:
                return tflops

        return 0

    def _fill_template(self, inp: SizingInput) -> openpyxl.Workbook:
        """
        Loads the template and populates the Inputs sheet's INPUT cells with
        values from SizingInput. Cell addresses match the layout produced by
        `llm_calc.inputs.build_inputs()` — keep these in sync if the source
        Inputs layout changes.

        GPU model, LLM model, and quantization are dropdown-driven (lookup
        from the Reference sheet) and are NOT pre-filled here — the user
        picks them in Excel after download. The numeric/scalar inputs ARE
        pre-filled so the recalculation matches what the web showed.
        """
        wb = openpyxl.load_workbook(self.template_path)
        # The new template has multiple sheets; the inputs live on "Inputs".
        ws = wb["Inputs"] if "Inputs" in wb.sheetnames else wb.active

        # ── Workload (4 segments; web sends 2: internal + external) ──
        # Segment columns: D=Внутренние, E=Внешние, F/G=spare segments.
        ws["D7"] = inp.internal_users
        ws["D8"] = inp.penetration_internal
        ws["D9"] = inp.concurrency_internal
        ws["D10"] = inp.sessions_per_user_J

        ws["E7"] = inp.external_users
        ws["E8"] = inp.penetration_external
        ws["E9"] = inp.concurrency_external
        ws["E10"] = inp.sessions_per_user_J

        for col in ("F", "G"):
            ws[f"{col}7"] = 0
            ws[f"{col}8"] = 0
            ws[f"{col}9"] = 0
            ws[f"{col}10"] = 0

        # ── Tokens (Section 2.2) ──
        ws["D15"] = inp.system_prompt_tokens_SP
        ws["D16"] = inp.user_prompt_tokens_Prp
        ws["D17"] = inp.answer_tokens_A
        ws["D18"] = inp.reasoning_tokens_MRT
        ws["D19"] = inp.dialog_turns

        # ── Hardware (Section 4) ──
        # GPU model (D25), quantization (D53), and model name (D35) are
        # left as their dropdown defaults — the user re-picks in Excel.
        ws["D29"] = inp.gpus_per_server
        ws["D30"] = inp.kavail
        ws["D31"] = inp.tp_multiplier_Z

        # ── Calibration coefficients (Section 3.1 / Е) ──
        ws["D55"] = inp.bytes_per_kv_state
        ws["D56"] = inp.emp_model
        ws["D57"] = inp.emp_kv
        ws["D58"] = inp.safe_margin

        # ── Compute calibration (Section 6 / Е.5) ──
        ws["D61"] = inp.eta_prefill
        ws["D62"] = inp.eta_decode
        ws["D64"] = inp.saturation_coeff_C
        ws["D65"] = inp.th_prefill_empir if inp.th_prefill_empir else 0
        ws["D66"] = inp.th_decode_empir if inp.th_decode_empir else 0
        # eta_cache (prefix-cache hit fraction) is on D67 — only written
        # when present in input; backend default in SizingInput is 0.0.
        if inp.eta_cache is not None:
            ws["D67"] = inp.eta_cache

        # ── SLA section (6.4 / 7) ──
        ws["D75"] = inp.rps_per_session_R
        ws["D76"] = inp.sla_reserve_KSLA
        ws["D77"] = inp.ttft_sla if inp.ttft_sla else 0
        ws["D78"] = inp.e2e_latency_sla if inp.e2e_latency_sla else 0

        return wb

    def generate(self, inp: SizingInput) -> io.BytesIO:
        """
        Генерирует заполненный Excel-файл и возвращает его как BytesIO-буфер.

        Raises:
            FileNotFoundError: если шаблон не найден.
            RuntimeError: при ошибке генерации.
        """
        if not self.template_exists:
            raise FileNotFoundError(f"Шаблон отчёта не найден: {self.template_path}")

        try:
            wb = self._fill_template(inp)
        except Exception as exc:
            logger.error("Ошибка при заполнении шаблона: %s", exc)
            raise RuntimeError(f"Ошибка генерации отчёта: {exc}") from exc

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def make_filename() -> str:
        """Формирует имя файла с текущей датой/временем."""
        return f"sizing_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
